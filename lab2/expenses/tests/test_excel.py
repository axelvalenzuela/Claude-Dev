import shutil
import tempfile
from io import BytesIO

from django.core.files.uploadedfile import SimpleUploadedFile
from django.test import TestCase, override_settings
from openpyxl import load_workbook

from accounts.models import User
from expenses.excel import build_report_workbook
from expenses.models import ExpenseReport, TravelDocument

MEDIA_ROOT = tempfile.mkdtemp(prefix="expense_reports_tests_")


@override_settings(MEDIA_ROOT=MEDIA_ROOT)
class ExcelReportTests(TestCase):
    @classmethod
    def tearDownClass(cls):
        super().tearDownClass()
        shutil.rmtree(MEDIA_ROOT, ignore_errors=True)

    def setUp(self):
        self.employee = User.objects.create_user(
            username="ana@example.com",
            email="ana@example.com",
            password="x",
            first_name="Ana Perez",
            department="Sales",
        )
        self.report = ExpenseReport.objects.create(user=self.employee, title="Trip to Mexico City")
        TravelDocument.objects.create(
            expense_report=self.report,
            file=SimpleUploadedFile("flight.pdf", b"x", content_type="application/pdf"),
            type=TravelDocument.DocType.FLIGHT,
            amount="2500.00",
            document_date="2026-08-01",
        )
        TravelDocument.objects.create(
            expense_report=self.report,
            file=SimpleUploadedFile("hotel.pdf", b"x", content_type="application/pdf"),
            type=TravelDocument.DocType.HOTEL,
            amount="1800.00",
            document_date="2026-08-02",
        )

    def test_generates_workbook_with_documents_and_total(self):
        workbook = build_report_workbook(self.report)

        buffer = BytesIO()
        workbook.save(buffer)
        buffer.seek(0)

        loaded = load_workbook(buffer)
        sheet = loaded.active

        values = [cell.value for row in sheet.iter_rows() for cell in row if cell.value is not None]
        self.assertIn("Ana Perez", values)
        self.assertIn("flight.pdf", values)
        self.assertIn("hotel.pdf", values)
        self.assertIn(4300.0, values)

    def test_includes_daily_breakdown_section(self):
        workbook = build_report_workbook(self.report)
        sheet = workbook.active

        values = [cell.value for row in sheet.iter_rows() for cell in row if cell.value is not None]
        self.assertTrue(any("Daily breakdown" in str(v) for v in values))

    def test_generates_workbook_without_documents(self):
        empty_report = ExpenseReport.objects.create(user=self.employee, title="No documents")
        workbook = build_report_workbook(empty_report)
        self.assertEqual(len(workbook.worksheets), 1)
