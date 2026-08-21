"""Builds the .xlsx expense report from an ExpenseReport, including the
per-day breakdown against the company's $60/day policy and, once approved,
the CEO approval clause."""
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill

from .policies import DAILY_LIMIT_USD, USD_MXN_RATE

HEADER_FILL = PatternFill("solid", fgColor="D9D9D9")
VIOLATION_FILL = PatternFill("solid", fgColor="F8D7DA")
# A flight or hotel charge routinely clears $60/day on its own — that's
# expected, not a policy question, so it gets a neutral informational
# color instead of the red used for an unexplained overage.
JUSTIFIED_FILL = PatternFill("solid", fgColor="D9E8F5")
# Same green the company's real Advance & Expense Report spreadsheet uses
# to highlight its own totals column.
USD_FILL = PatternFill("solid", fgColor="C6EFCE")


def build_report_workbook(report) -> Workbook:
    wb = Workbook()
    sheet = wb.active
    sheet.title = "Expense report"

    sheet["A1"] = "Travel expense report"
    sheet["A1"].font = Font(bold=True, size=14)

    trip_range = report.trip_date_range
    trip_range_text = f"{trip_range[0]:%Y-%m-%d} to {trip_range[1]:%Y-%m-%d}" if trip_range else "—"

    info_rows = 2
    for label, value in [
        ("Employee", report.user.get_full_name() or report.user.email),
        ("Employee #", report.user.employee_number or "—"),
        ("Department", report.user.department),
        ("Supervisor", report.supervisor_name),
        ("Title", report.title),
        ("Trip dates", trip_range_text),
        ("Status", report.get_status_display()),
        ("Created", report.created_at.strftime("%Y-%m-%d %H:%M")),
    ]:
        info_rows += 1
        sheet[f"A{info_rows}"], sheet[f"B{info_rows}"] = label, value

    if report.reviewed_at:
        info_rows += 1
        sheet[f"A{info_rows}"], sheet[f"B{info_rows}"] = "Reviewed", report.reviewed_at.strftime("%Y-%m-%d %H:%M")
        info_rows += 1
        sheet[f"A{info_rows}"], sheet[f"B{info_rows}"] = "Review note", report.review_note
        if report.approval_clause:
            info_rows += 1
            sheet[f"A{info_rows}"], sheet[f"B{info_rows}"] = "Approval clause", report.approval_clause

    # Column layout mirrors the company's real Advance & Expense Report
    # spreadsheet (Invoice date / Description / Vendor legal name / FX
    # Rate / ... ), with "Amount (USD)" replacing that sheet's "Total MXN"
    # column since this app's policy and totals are USD-based.
    header_row = info_rows + 2
    headers = [
        "#", "Invoice date", "Description", "Vendor legal name",
        "Expensed amount\nin foreign currency", "Currency", "FX Rate",
        "Backup", "File", "Amount (USD)",
    ]
    for col, text in enumerate(headers, start=1):
        cell = sheet.cell(row=header_row, column=col, value=text)
        cell.font = Font(bold=True)
        cell.fill = USD_FILL if text == "Amount (USD)" else HEADER_FILL
        cell.alignment = cell.alignment.copy(wrap_text=True)

    row = header_row + 1
    total_usd = 0
    for index, doc in enumerate(report.documents.order_by("type", "document_date"), start=1):
        sheet.cell(row=row, column=1, value=index)
        sheet.cell(row=row, column=2, value=doc.document_date.strftime("%Y-%m-%d"))
        sheet.cell(row=row, column=3, value=doc.get_type_display())
        sheet.cell(row=row, column=4, value=doc.vendor_name or "—")
        sheet.cell(row=row, column=5, value=float(doc.amount))
        sheet.cell(row=row, column=6, value=doc.currency)
        sheet.cell(row=row, column=7, value=float(1) if doc.currency == "USD" else float(USD_MXN_RATE))
        sheet.cell(row=row, column=8, value=doc.get_backup_type_display())
        sheet.cell(row=row, column=9, value=doc.file_name)
        usd_cell = sheet.cell(row=row, column=10, value=float(doc.amount_usd))
        usd_cell.fill = USD_FILL
        total_usd += doc.amount_usd
        row += 1

    sheet.cell(row=row, column=9, value="Total (USD)").font = Font(bold=True)
    total_cell = sheet.cell(row=row, column=10, value=float(total_usd))
    total_cell.font = Font(bold=True)
    total_cell.fill = USD_FILL
    row += 2

    # Daily breakdown against the $60/day policy — always compared in USD
    # (see ExpenseReport.daily_totals / TravelDocument.amount_usd), since a
    # peso figure compared directly against a dollar limit would be
    # meaningless.
    sheet.cell(row=row, column=1, value=f"Daily breakdown (policy limit: ${DAILY_LIMIT_USD}/day)").font = Font(bold=True)
    row += 1
    daily_header_row = row
    for col, text in enumerate(["Date", "Daily total", "Daily total (USD)", "Over policy limit?"], start=1):
        cell = sheet.cell(row=daily_header_row, column=col, value=text)
        cell.font = Font(bold=True)
        cell.fill = HEADER_FILL
    row += 1
    for day in report.daily_totals():
        sheet.cell(row=row, column=1, value=day["date"].strftime("%Y-%m-%d"))
        sheet.cell(row=row, column=2, value=float(day["total"]))
        sheet.cell(row=row, column=3, value=float(day["total_usd"]))
        if day["over_limit"]:
            flag_value, fill = "Yes", VIOLATION_FILL
        elif day["has_flight_or_hotel"]:
            flag_value, fill = "Flight/Hotel", JUSTIFIED_FILL
        else:
            flag_value, fill = "No", None
        sheet.cell(row=row, column=4, value=flag_value)
        if fill:
            for col in (1, 2, 3, 4):
                sheet.cell(row=row, column=col).fill = fill
        row += 1

    for col in "ABCDEFGHIJ":
        sheet.column_dimensions[col].width = 22

    return wb
